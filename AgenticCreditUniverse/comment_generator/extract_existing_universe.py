"""기존 유니버스 엑셀에서 Stage 2 (judgment-reviewer) 입력 데이터 추출.

엑셀 컬럼(1-based):
  1=발행기관, 5=25.2H 신용등급, 6=25.2H 등급전망,
  7=26.1H 신용등급, 8=26.1H 등급전망,
  9=25.2H 유니버스, 10=26.1H 유니버스

Stage 2 가 풀 단위 형평성·안정성 가드레일을 적용하려면:
  - 검토 종목별 직전 반기(25.2H) 분류 (`prior_lookup`)
  - 기존 유니버스 전체의 (등급, 전망, 25.2H 분류) 컨텍스트 (`existing_universe`)

분류 값은 O/△/X 만 표준 — 그 외 텍스트는 None 으로 normalize.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


# 엑셀 컬럼 인덱스 (1-based, ws.cell(row, col) 호출용)
_COL_NAME = 1
_COL_GRADE_25_2H = 5
_COL_OUTLOOK_25_2H = 6
_COL_GRADE_26_1H = 7
_COL_OUTLOOK_26_1H = 8
_COL_UNIVERSE_25_2H = 9
_COL_UNIVERSE_26_1H = 10

# 사용자 엑셀에 들어올 가능성 있는 분류 표기 → 표준 (O/△/X)
_CLASS_MAP = {
    "O": "O", "o": "O", "○": "O", "동그라미": "O",
    "△": "△", "▲": "△", "세모": "△",
    "X": "X", "x": "X", "✕": "X", "엑스": "X",
}


def _normalize_class(v: object) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    return _CLASS_MAP.get(s)


def _normalize_text(v: object) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def extract(xlsx_path: Path) -> dict:
    """기존 유니버스 엑셀에서 Stage 2 입력 데이터 추출.

    Returns:
      {
        "existing_universe": [
          {
            "name": str,
            "grade_25_2h": str|None,
            "outlook_25_2h": str|None,
            "grade_26_1h": str|None,
            "outlook_26_1h": str|None,
            "universe_25_2h": str|None,  # O/△/X
          },
          ...
        ],
        "prior_lookup": {name: universe_25_2h_or_None, ...},
      }

    엑셀 첫 행은 헤더로 가정. 발행기관 셀이 비어있으면 행 스킵.
    동일 발행기관 다중 행은 첫 행만 사용(머지 로직과 일관).
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    rows: list[dict] = []
    seen: set[str] = set()
    for r in range(2, ws.max_row + 1):
        name = _normalize_text(ws.cell(r, _COL_NAME).value)
        if not name or name in seen:
            continue
        seen.add(name)
        rows.append(
            {
                "name": name,
                "grade_25_2h": _normalize_text(ws.cell(r, _COL_GRADE_25_2H).value),
                "outlook_25_2h": _normalize_text(ws.cell(r, _COL_OUTLOOK_25_2H).value),
                "grade_26_1h": _normalize_text(ws.cell(r, _COL_GRADE_26_1H).value),
                "outlook_26_1h": _normalize_text(ws.cell(r, _COL_OUTLOOK_26_1H).value),
                "universe_25_2h": _normalize_class(ws.cell(r, _COL_UNIVERSE_25_2H).value),
            }
        )
    wb.close()

    prior_lookup = {row["name"]: row["universe_25_2h"] for row in rows}
    return {"existing_universe": rows, "prior_lookup": prior_lookup}


if __name__ == "__main__":
    import argparse
    import json
    import sys

    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("xlsx", type=Path, help="기존 유니버스 엑셀 경로")
    p.add_argument(
        "--output",
        type=Path,
        default=None,
        help="결과 JSON 저장 경로 (생략 시 stdout)",
    )
    args = p.parse_args()

    if not args.xlsx.exists():
        print(f"[ERROR] xlsx not found: {args.xlsx}", file=sys.stderr)
        sys.exit(2)

    result = extract(args.xlsx)
    out = json.dumps(result, ensure_ascii=False, indent=2)
    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(out, encoding="utf-8")
        print(f"saved: {args.output}", file=sys.stderr)
    else:
        print(out)
