"""
Build _workspace/index.sqlite from:
  - output/26.1H 유니버스.xlsx  (998 rows × 18 cols, formulas + evaluated values)
  - _workspace/master/master.json
  - _workspace/comments/{slug}.json
  - _workspace/judgment/stage2_review.json
  - _workspace/nice/{slug}/, dart/{slug}/, news/{slug}/{key}/

Run: .venv/bin/python web/backend/scripts/build_index.py
Idempotent — drops and recreates schema each run.
"""
from __future__ import annotations

import json
import os
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl

# 경로는 환경변수 우선. 로컬 개발은 web/.env.example 의 WORKSPACE_DIR/EXCEL_PATH
# 를 백엔드 .env 에 채워서 실행, 컨테이너 배포는 Dockerfile 의 ENV 가 동일 변수 주입.
# default 계산은 lazy — 컨테이너 안에서는 parents[4] 가 IndexError 라 env 가 있을 때만 default 우회.
def _ws_default() -> Path:
    return Path(__file__).resolve().parents[4] / "_workspace"


def _xlsx_default() -> Path:
    return Path(__file__).resolve().parents[4] / "output" / "26.1H 유니버스.xlsx"


_WS_ENV = os.environ.get("WORKSPACE_DIR")
_XLSX_ENV = os.environ.get("EXCEL_PATH")
WS = Path(_WS_ENV) if _WS_ENV else _ws_default()
XLSX = Path(_XLSX_ENV) if _XLSX_ENV else _xlsx_default()
DB = WS / "index.sqlite"
SCHEMA = Path(__file__).resolve().parent.parent / "app" / "db_schema.sql"

DEFAULT_PERIOD = {
    "current": "26.1H",
    "previous": "25.2H",
    "previous_previous": "25.1H",
    "history": ["25.1H", "25.2H", "26.1H"],
}


def _norm(s: str | None) -> str:
    if not s:
        return ""
    return "".join(str(s).lower().split())


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def _mtime_iso(p: Path) -> str | None:
    try:
        return datetime.fromtimestamp(p.stat().st_mtime, tz=timezone.utc).isoformat(timespec="seconds")
    except FileNotFoundError:
        return None


def load_master() -> dict[str, dict[str, Any]]:
    path = WS / "master" / "master.json"
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8")).get("companies", {})


def load_watch_industries() -> set[str]:
    """유의업종 카테고리 SSOT (`_workspace/master/watch_industries.json`).

    사용자가 매 반기 갱신. industry 가 categories 에 정확히 매칭되면
    industry_2026 = "O" 자동 부여.
    """
    path = WS / "master" / "watch_industries.json"
    if not path.exists():
        return set()
    data = json.loads(path.read_text(encoding="utf-8"))
    return {str(c).strip() for c in data.get("categories", []) if str(c).strip()}


def build_alias_map(master: dict[str, dict[str, Any]]) -> dict[str, str]:
    """norm(alias) -> slug. issuer 매칭 시 _norm 후 lookup."""
    m: dict[str, str] = {}
    for slug, info in master.items():
        candidates = {slug, info.get("official_name") or "", *(info.get("aliases") or [])}
        for c in candidates:
            if c:
                m[_norm(c)] = slug
    return m


def latest_mtime_for_slug(slug: str) -> str | None:
    paths: list[Path] = []
    cmt = WS / "comments" / f"{slug}.json"
    if cmt.exists():
        paths.append(cmt)
    nice_dir = WS / "nice" / slug
    if nice_dir.exists():
        paths.extend(p for p in nice_dir.glob("*.json"))
    dart_meta = WS / "dart" / slug / "metadata.json"
    if dart_meta.exists():
        paths.append(dart_meta)
    news_dir = WS / "news" / slug
    if news_dir.exists():
        paths.extend(p for p in news_dir.rglob("metadata.json"))
    if not paths:
        return None
    latest = max(paths, key=lambda p: p.stat().st_mtime)
    return _mtime_iso(latest)


def open_excel_dual() -> tuple[openpyxl.worksheet.worksheet.Worksheet, openpyxl.worksheet.worksheet.Worksheet]:
    """formulas (data_only=False) + evaluated values (data_only=True)."""
    wb_f = openpyxl.load_workbook(XLSX, data_only=False, read_only=True)
    wb_v = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
    return wb_f.active, wb_v.active


COLS = (
    "issuer", "request_class", "industry_2026", "industry",
    "rating_prev", "watch_prev", "rating_curr", "watch_curr",
    "universe_prev", "universe_curr_ai", "manager",
    "comment_prev", "comment_curr",
    "_movement_formula",   # Col 14 — formula in ws_f, evaluated in ws_v
    "group_name", "ai_judgment", "ai_rationale", "reviewer_final",
)


def cell_str(v: Any) -> str | None:
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    return str(v)


_UNIVERSE_RANK = {"O": 3, "△": 2, "X": 1}


def _compute_movement(prev: str | None, decision: str | None) -> str | None:
    """전기 유니버스 ↔ 당기 결정(심사역 최종 우선, 없으면 AI 판단) 비교.

    Excel Col 14 의 수식은 openpyxl 가 평가하지 않으므로 (Excel 가 한 번도 열린 적 없으면)
    런타임에 직접 계산한다. 또한 사용자가 Col 18 (심사역 최종) 을 채우면 이 값이
    의견변동에 반영되어야 함 (기존 수식은 J열=AI 만 보고 있었음).
    """
    if not prev or not decision:
        return None
    p = _UNIVERSE_RANK.get(prev.strip())
    c = _UNIVERSE_RANK.get(decision.strip())
    if p is None or c is None:
        return None
    if c > p:
        return "▲"
    if c < p:
        return "▽"
    return "-"


def build() -> None:
    if not XLSX.exists():
        raise FileNotFoundError(f"Excel not found: {XLSX}")
    if not SCHEMA.exists():
        raise FileNotFoundError(f"Schema not found: {SCHEMA}")

    master = load_master()
    alias_map = build_alias_map(master)
    watch_categories = load_watch_industries()

    stage2_path = WS / "judgment" / "stage2_review.json"
    _ = json.loads(stage2_path.read_text(encoding="utf-8")) if stage2_path.exists() else {}
    review_path = WS / "review_status.json"
    review = json.loads(review_path.read_text(encoding="utf-8")) if review_path.exists() else {}

    ws_f, ws_v = open_excel_dual()

    DB.parent.mkdir(parents=True, exist_ok=True)
    if DB.exists():
        DB.unlink()
    con = sqlite3.connect(DB)
    con.executescript(SCHEMA.read_text(encoding="utf-8"))

    rows: list[tuple] = []
    now = _now_iso()
    for r in range(2, ws_f.max_row + 1):
        issuer = cell_str(ws_f.cell(r, 1).value)
        if not issuer:
            continue

        slug = alias_map.get(_norm(issuer))
        unresolved = 0 if slug else 1
        if not slug:
            slug = f"unresolved::{r}::{_norm(issuer)[:32]}"

        row_vals: dict[str, Any] = {}
        for i, key in enumerate(COLS, start=1):
            row_vals[key] = cell_str(ws_f.cell(r, i).value)
        # 의견변동 — 심사역 최종(있으면) > AI 판단 우선순위로 동적 계산.
        # Excel C14 수식은 openpyxl 평가 미지원으로 항상 None → Python 계산이 SSOT.
        decision = row_vals.get("reviewer_final") or row_vals.get("universe_curr_ai")
        row_vals["movement"] = _compute_movement(row_vals.get("universe_prev"), decision)

        # 유의업종 자동 판정 — 사용자가 엑셀 Col 3 에 명시 'O' 입력한 경우는 보존.
        # 빈 셀에 한해 industry ∈ watch_categories 면 'O' 부여.
        if not row_vals.get("industry_2026"):
            ind = (row_vals.get("industry") or "").strip()
            if ind and ind in watch_categories:
                row_vals["industry_2026"] = "O"

        info = master.get(slug, {}) if not unresolved else {}

        rows.append((
            slug, r,
            row_vals["issuer"], row_vals["request_class"], row_vals["industry_2026"], row_vals["industry"],
            row_vals["rating_prev"], row_vals["watch_prev"], row_vals["rating_curr"], row_vals["watch_curr"],
            row_vals["universe_prev"], row_vals["universe_curr_ai"], row_vals["manager"],
            row_vals["comment_prev"], row_vals["comment_curr"], row_vals["movement"],
            row_vals["group_name"], row_vals["ai_judgment"], row_vals["ai_rationale"], row_vals["reviewer_final"],
            info.get("stock_code"), info.get("corp_code"), info.get("cmp_cd"),
            info.get("official_name"), info.get("group"),
            unresolved,
            None if unresolved else latest_mtime_for_slug(slug),
            now,
        ))

    con.executemany(
        """INSERT INTO companies (
            slug, excel_row,
            issuer, request_class, industry_2026, industry,
            rating_prev, watch_prev, rating_curr, watch_curr,
            universe_prev, universe_curr_ai, manager,
            comment_prev, comment_curr, movement,
            group_name, ai_judgment, ai_rationale, reviewer_final,
            stock_code, corp_code, cmp_cd, official_name, group_master,
            unresolved, last_updated_utc, created_at
        ) VALUES (?,?,?,?,?,?, ?,?,?,?, ?,?,?, ?,?,?, ?,?,?,?, ?,?,?,?,?, ?,?,?)""",
        rows,
    )

    # review_status mirror
    for slug, st in review.items():
        con.execute(
            """INSERT INTO review_status (slug, status, universe, agree_with_ai, note, reviewed_by, reviewed_at)
               VALUES (?,?,?,?,?,?,?)""",
            (
                slug,
                st.get("status", "none"),
                st.get("universe"),
                int(bool(st.get("agree_with_ai"))) if st.get("agree_with_ai") is not None else None,
                st.get("note"),
                st.get("reviewed_by"),
                st.get("reviewed_at"),
            ),
        )

    # period_config (write fallback to disk if missing)
    pc_path = WS / "period_config.json"
    if pc_path.exists():
        pc = json.loads(pc_path.read_text(encoding="utf-8"))
    else:
        pc = DEFAULT_PERIOD
        pc_path.write_text(json.dumps(pc, ensure_ascii=False, indent=2), encoding="utf-8")
    for k, v in pc.items():
        con.execute(
            "INSERT INTO period_config (k, v) VALUES (?, ?)",
            (k, json.dumps(v, ensure_ascii=False)),
        )

    con.commit()

    total = con.execute("SELECT COUNT(*) FROM companies").fetchone()[0]
    resolved = con.execute("SELECT COUNT(*) FROM companies WHERE unresolved=0").fetchone()[0]
    con.close()
    print(f"[index] built {DB}  total={total}  resolved={resolved}  unresolved={total - resolved}")


if __name__ == "__main__":
    build()
