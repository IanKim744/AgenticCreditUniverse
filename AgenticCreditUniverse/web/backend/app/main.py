from __future__ import annotations

import json
import sqlite3
import subprocess
import sys
from contextlib import asynccontextmanager
from io import BytesIO
from pathlib import Path
from urllib.parse import quote

import openpyxl
from openpyxl.styles import Alignment, Font
from fastapi import Depends, FastAPI
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse

from .auth import require_session
from .deps import get_db
from .routers import auth as auth_router
from .routers import companies as companies_router
from .routers import review as review_router


def _rebuild_index() -> None:
    here = Path(__file__).resolve().parent
    script = here.parent / "scripts" / "build_index.py"
    if not script.exists():
        return
    subprocess.run([sys.executable, str(script)], check=True)


@asynccontextmanager
async def lifespan(app: FastAPI):
    _rebuild_index()
    yield


app = FastAPI(title="크레딧 유니버스 API", lifespan=lifespan)

# 개발 편의용 CORS — 프록시(rewrites) 사용 시 same-origin이지만 직호출 대비.
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

app.include_router(auth_router.router)
app.include_router(companies_router.router)
app.include_router(review_router.router)


@app.get("/api/healthz")
def healthz() -> dict[str, str]:
    return {"status": "ok"}


_MATRIX_HEADERS: tuple[str, ...] = (
    "발행기관",
    "그룹사",
    "업종",
    "유의업종",
    "전기 등급",
    "당기 등급",
    "전기 유니버스",
    "당기 유니버스 (AI)",
    "심사역 최종 판단",
    "의견변동",
    "당기 코멘트",
    "담당",
    "검수여부",
    "마지막 갱신",
)
# 매트릭스 화면(MatrixTable) 의 컬럼 너비를 엑셀에서도 비슷하게.
_MATRIX_COL_WIDTHS: tuple[int, ...] = (22, 16, 18, 8, 16, 16, 12, 14, 14, 10, 80, 10, 10, 22)


def _grade_cell(rating: str | None, watch: str | None) -> str:
    r = (rating or "").strip()
    w = (watch or "").strip()
    if r and w:
        return f"{r} ({w})"
    return r


def _current_period(con: sqlite3.Connection) -> str:
    row = con.execute("SELECT v FROM period_config WHERE k='current'").fetchone()
    if not row:
        return ""
    try:
        return json.loads(row["v"])
    except (json.JSONDecodeError, TypeError):
        return str(row["v"])


@app.get("/api/export.xlsx")
def export_matrix_excel(
    _sess: dict = Depends(require_session),
    con: sqlite3.Connection = Depends(get_db),
) -> StreamingResponse:
    """매트릭스 화면(MatrixTable) 과 동일한 14-컬럼 스키마로 xlsx 생성.

    - 사이드바/검색 필터는 무시 — 항상 전체 행.
    - `comment_curr` 전문, 등급+전망 결합 셀.
    - SSOT: SQLite index (`/api/companies` 와 동일 소스).
    """
    review_rows = con.execute(
        "SELECT slug, status FROM review_status"
    ).fetchall()
    review_status_map = {r["slug"]: r["status"] for r in review_rows}

    companies = con.execute(
        "SELECT * FROM companies ORDER BY excel_row"
    ).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "매트릭스"

    ws.append(list(_MATRIX_HEADERS))
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
    ws.freeze_panes = "A2"

    wrap_top = Alignment(wrap_text=True, vertical="top")
    for r in companies:
        review_done = (
            review_status_map.get(r["slug"]) == "done"
            or bool(r["reviewer_final"])
        )
        ws.append([
            r["issuer"],
            r["group_name"] or r["group_master"],
            r["industry"],
            r["industry_2026"],
            _grade_cell(r["rating_prev"], r["watch_prev"]),
            _grade_cell(r["rating_curr"], r["watch_curr"]),
            r["universe_prev"],
            r["universe_curr_ai"] or r["ai_judgment"],
            r["reviewer_final"],
            r["movement"],
            r["comment_curr"],
            r["manager"],
            "완료" if review_done else "미검수",
            r["last_updated_utc"],
        ])
        # 코멘트 셀(11열) 자동 줄바꿈
        ws.cell(row=ws.max_row, column=11).alignment = wrap_top

    for idx, width in enumerate(_MATRIX_COL_WIDTHS, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    period = _current_period(con) or "current"
    filename = f"유니버스_매트릭스_{period}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                f"attachment; filename=\"matrix.xlsx\"; "
                f"filename*=UTF-8''{quote(filename)}"
            )
        },
    )
