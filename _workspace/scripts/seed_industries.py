"""
26.1H 유니버스 18종목 업종 백필 + 유의업종 카테고리 SSOT 생성.

- _workspace/master/master.json: companies[*].industry 채움
- output/26.1H 유니버스.xlsx: Col 4 (업종) 채움
- _workspace/master/watch_industries.json: 유의업종 카테고리 리스트(7개) 신규

업종 분류는 신평사(한기평/한신평) Industry Credit Outlook 수준 카테고리.
사용자 검수 완료(2026-04-26).
"""
from __future__ import annotations
import json
from datetime import date
from pathlib import Path

import openpyxl

ROOT = Path("/Users/kim-youngbyeok/AgenticCreditUniverse/AgenticCreditUniverse")
MASTER = ROOT / "_workspace" / "master" / "master.json"
WATCH = ROOT / "_workspace" / "master" / "watch_industries.json"
XLSX = ROOT / "output" / "26.1H 유니버스.xlsx"

# 발행기관(엑셀 Col 1 == master.json key) → 업종.
# 사용자 검수 완료 (2026-04-26). 신평사 표준 분류 수준.
INDUSTRY_MAP: dict[str, str] = {
    "다우기술": "IT서비스",
    "대한해운": "해운",
    "롯데물산": "부동산",
    "보령(구. 보령제약)": "제약",
    "부산롯데호텔": "호텔/레저",
    "씨제이프레시웨이": "음식료",
    "알씨아이파이낸셜서비스코리아": "캐피탈",
    "에스케이디스커버리": "석유화학",
    "에스케이스페셜티(구. 에스케이머티리얼즈)": "화학",
    "에스케이실트론": "반도체",
    "에스케이온": "2차전지",
    "에이치라인해운": "해운",
    "한솔제지": "제지",
    "한솔케미칼": "화학",
    "한솔테크닉스": "전기전자",
    "한일시멘트": "시멘트",
    "한화리츠": "리츠",
    "현대로템": "기계",
    # 엑셀에 추가로 존재하는 4종목 — 같은 분류 기준 적용.
    # 사용자가 이미 채운 셀(현대비앤지스틸=철강, 에스케이씨=석유화학)은 그대로 두면
    # update_excel() 에서 동일값으로 덮어써도 무해. 기존 값 보존 의미로 매핑에 명시.
    "현대리바트": "가구",
    "현대비앤지스틸": "철강",
    "에스케이씨": "석유화학",
    "에스케이네트웍스서비스": "IT서비스",
}

# 26.1H 유의업종 카테고리 — 사용자 운영(매 반기 갱신).
WATCH_INDUSTRIES: list[str] = [
    "2차전지",
    "석유화학",
    "철강",
    "상영관",
    "건설",
    "저축은행",
    "부동산신탁",
]


def update_master() -> None:
    data = json.loads(MASTER.read_text(encoding="utf-8"))
    companies = data["companies"]
    missing: list[str] = []
    for issuer, ind in INDUSTRY_MAP.items():
        if issuer not in companies:
            missing.append(issuer)
            continue
        companies[issuer]["industry"] = ind
    data["updated_at"] = date.today().isoformat()
    MASTER.write_text(
        json.dumps(data, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"[master] updated {len(INDUSTRY_MAP) - len(missing)} entries")
    if missing:
        print(f"[master] WARN missing: {missing}")


def update_excel() -> None:
    wb = openpyxl.load_workbook(XLSX)
    ws = wb.active
    updated = 0
    skipped: list[str] = []
    for r in range(2, ws.max_row + 1):
        issuer = ws.cell(r, 1).value
        if not issuer:
            continue
        issuer = str(issuer).strip()
        if issuer in INDUSTRY_MAP:
            ws.cell(r, 4).value = INDUSTRY_MAP[issuer]
            updated += 1
        else:
            skipped.append(issuer)
    wb.save(XLSX)
    print(f"[excel] updated Col 4 (업종) for {updated} rows")
    if skipped:
        print(f"[excel] note unmapped issuers: {skipped}")


def write_watch_industries() -> None:
    payload = {
        "version": "1.0",
        "updated_at": date.today().isoformat(),
        "period": "26.1H",
        "categories": WATCH_INDUSTRIES,
        "notes": "사용자 운영. 매 반기 갱신 시 categories 리스트만 수정하면 build_index.py가 자동 ○ 판정에 반영.",
    }
    WATCH.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"[watch] wrote {len(WATCH_INDUSTRIES)} categories → {WATCH.name}")


if __name__ == "__main__":
    update_master()
    update_excel()
    write_watch_industries()
