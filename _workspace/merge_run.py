"""excel-merge SKILL.md 의사코드 실행 스크립트 (18개 컬럼, 자동 갱신 7개).

base = output/26.1H 자체 (in-place 갱신). 사용자 수기 작업이 계속 누적되도록.
첫 빌드 시점에는 legacy 26.1Q 를 OUTPUT 위치로 한 번 복사한 뒤 마이그레이션 스크립트
(_workspace/migrate_18cols.py) 로 18개 컬럼 만든 상태가 사전 조건.

자동 갱신: col 13 (26.1H 코멘트), col 14 (의견변동 수식),
          col 16 (AI 판단), col 17 (AI 판단 사유)
수기 보존: col 1~6, 9, 11, 12, 15, 18 — 절대 미갱신
"""
from __future__ import annotations

import json
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment


ROOT = Path("/Users/kim-youngbyeok/AgenticCreditUniverse/AgenticCreditUniverse")
COMMENTS = ROOT / "_workspace/comments"
JUDGMENT = ROOT / "_workspace/judgment/stage2_review.json"
OUTPUT = ROOT / "output/26.1H 유니버스.xlsx"
BACKUP_DIR = ROOT / "output/backup"

EXPECTED_18 = [
    "발행기관", "현업 요청 분류", "26년\n유의업종", "업종",
    "25.2H 신용등급", "25.2H\n등급전망", "26.1H 신용등급",
    "26.1H\n등급전망", "25.2H 유니버스", "26.1H 유니버스",
    "26.1H\n담당", "25.2H 검토 코멘트", "26.1H 검토 코멘트",
    "26.1H 유니버스 의견변동", "그룹사",
    "AI 판단", "AI 판단 사유", "심사역 최종 판단",
]


def _opinion_change_formula(r: int) -> str:
    """col 14 의견변동 수식. I=col9 (25.2H), J=col10 (26.1H). O>△>X 순위 비교.

    같으면 "-", 26.1H 더 높으면 ▲, 더 낮으면 ▽, 한쪽 빈 셀이면 빈 문자열.
    """
    return (
        f'=IF(OR(I{r}="",J{r}=""),"",'
        f'IF(I{r}=J{r},"-",'
        f'IF((IF(J{r}="O",3,IF(J{r}="△",2,IF(J{r}="X",1,0))))>'
        f'(IF(I{r}="O",3,IF(I{r}="△",2,IF(I{r}="X",1,0)))),"▲","▽")))'
    )


def main() -> None:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    shutil.copy(OUTPUT, BACKUP_DIR / f"{ts}_pre-merge_26.1H.xlsx")

    wb = load_workbook(OUTPUT)  # in-place 갱신 (사용자 수기 셀 자동 보존)
    ws = wb.active

    headers = [c.value for c in ws[1]]
    for i, exp in enumerate(EXPECTED_18):
        if headers[i] != exp:
            raise ValueError(
                f"헤더 col{i+1} 불일치: {headers[i]!r} vs expected {exp!r}. "
                "마이그레이션(_workspace/migrate_18cols.py) 먼저 실행."
            )

    stage2 = json.loads(JUDGMENT.read_text(encoding="utf-8"))
    decisions = stage2.get("decisions", {})

    # 데이터 행 끝
    last_row = 1
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value:
            last_row = r

    name_to_row: dict[str, int] = {}
    for r in range(2, last_row + 1):
        name = ws.cell(r, 1).value
        if name:
            name_to_row.setdefault(name, r)

    report = {"updated": [], "added": [], "missing_decisions": []}
    rationale_align = Alignment(wrap_text=True, vertical="top", horizontal="left")

    for cmt_file in sorted(COMMENTS.glob("*.json")):
        rec = json.loads(cmt_file.read_text(encoding="utf-8"))
        name = rec["company"]
        comment = rec["comment"]
        dec = decisions.get(name, {})
        ai = dec.get("final")
        rationale = dec.get("rationale", "")

        if name in name_to_row:
            r = name_to_row[name]
            ws.cell(r, 13).value = comment
            if ai:
                ws.cell(r, 16).value = ai
            if rationale:
                cell = ws.cell(r, 17)
                cell.value = rationale
                cell.alignment = rationale_align
            report["updated"].append({"company": name, "row": r, "ai": ai, "len": len(comment)})
        else:
            new_r = last_row + 1
            ws.cell(new_r, 1).value = name
            ws.cell(new_r, 13).value = comment
            if ai:
                ws.cell(new_r, 16).value = ai
            if rationale:
                cell = ws.cell(new_r, 17)
                cell.value = rationale
                cell.alignment = rationale_align
            last_row = new_r
            name_to_row[name] = new_r
            report["added"].append({"company": name, "row": new_r, "ai": ai, "len": len(comment)})

    # col 14 의견변동 수식: 모든 데이터 행에 일괄 적용 (수식이라 안전)
    for r in range(2, last_row + 1):
        ws.cell(r, 14).value = _opinion_change_formula(r)

    produced = {f.stem for f in COMMENTS.glob("*.json")}
    for name in decisions.keys() - produced:
        report["missing_decisions"].append({"company": name, "reason": "comments 산출물 없음"})

    wb.save(OUTPUT)

    merge_report_path = ROOT / "_workspace/merge/report.json"
    merge_report_path.parent.mkdir(parents=True, exist_ok=True)
    merge_report_path.write_text(
        json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
