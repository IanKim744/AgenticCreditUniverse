"""1회 실행 마이그레이션: 17개 컬럼 → 18개 컬럼 + 의견변동 수식 + 삼성FN리츠 삭제.

수행:
  1) 백업
  2) col 17 (심사역 최종 판단) → col 18 로 이동 (헤더 + 스타일)
  3) col 17 새 헤더 = "AI 판단 사유"
  4) col 17 폭 50, wrap_text=True
  5) col 17 데이터 채움 (stage2_review.json rationale)
  6) col 14 r=2..N에 의견변동 수식 일괄
  7) 삼성FN리츠 행 삭제
  8) 저장
"""
from __future__ import annotations

import json
import shutil
from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment


ROOT = Path("/Users/kim-youngbyeok/AgenticCreditUniverse/AgenticCreditUniverse")
TARGET = ROOT / "output/26.1H 유니버스.xlsx"
JUDGMENT = ROOT / "_workspace/judgment/stage2_review.json"
BACKUP_DIR = ROOT / "output/backup"


def main() -> None:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    shutil.copy(TARGET, BACKUP_DIR / f"{ts}_pre-f-changes_26.1H.xlsx")

    wb = load_workbook(TARGET)
    ws = wb.active

    # 데이터 행 끝
    last_row = 1
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value:
            last_row = r

    # 1) col 17 → col 18 헤더 이동 + 스타일 복사
    src = ws.cell(1, 17)
    dst = ws.cell(1, 18)
    dst.value = src.value
    dst.font = copy(src.font)
    dst.alignment = copy(src.alignment)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)

    # 2) col 17 새 헤더 = "AI 판단 사유" + col 16 스타일 복사
    ref = ws.cell(1, 16)
    new = ws.cell(1, 17)
    new.value = "AI 판단 사유"
    new.font = copy(ref.font)
    new.alignment = copy(ref.alignment)
    new.fill = copy(ref.fill)
    new.border = copy(ref.border)

    # 3) col 17 폭 + wrap_text
    ws.column_dimensions["Q"].width = 50
    # col 18 폭은 col 16 기준
    ws.column_dimensions["R"].width = 18

    # 4) col 17 데이터 채움 (stage2 rationale)
    stage2 = json.loads(JUDGMENT.read_text(encoding="utf-8"))
    decisions = stage2.get("decisions", {})
    filled = []
    for r in range(2, last_row + 1):
        name = ws.cell(r, 1).value
        if name and name in decisions:
            cell = ws.cell(r, 17)
            cell.value = decisions[name].get("rationale", "")
            cell.alignment = Alignment(
                wrap_text=True, vertical="top", horizontal="left"
            )
            filled.append((r, name))

    # 5) col 14 의견변동 수식 일괄
    formula_count = 0
    for r in range(2, last_row + 1):
        f = (
            f'=IF(OR(I{r}="",J{r}=""),"",'
            f'IF(I{r}=J{r},"-",'
            f'IF((IF(J{r}="O",3,IF(J{r}="△",2,IF(J{r}="X",1,0))))>'
            f'(IF(I{r}="O",3,IF(I{r}="△",2,IF(I{r}="X",1,0)))),"▲","▽")))'
        )
        ws.cell(r, 14).value = f
        formula_count += 1

    # 6) 삼성FN리츠 행 삭제 (스캔하여 동적으로 찾음)
    deleted_rows = []
    # last_row 부터 역순으로 스캔하여 삭제 (인덱스 안정성)
    for r in range(last_row, 1, -1):
        name = ws.cell(r, 1).value
        if name and "삼성FN리츠" in str(name):
            ws.delete_rows(r, 1)
            deleted_rows.append((r, name))

    wb.save(TARGET)

    print(json.dumps(
        {
            "backup": str(BACKUP_DIR / f"{ts}_pre-f-changes_26.1H.xlsx"),
            "filled_rationale": filled,
            "formula_rows": formula_count,
            "deleted_rows": deleted_rows,
        },
        ensure_ascii=False,
        indent=2,
    ))


if __name__ == "__main__":
    main()
